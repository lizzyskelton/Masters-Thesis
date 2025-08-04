# Finding all values within a fixed 2m radius of a central point 

import pandas as pd
from geopy.distance import geodesic
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

def find_points_within_radius(csv_file, target_lat, target_lon, radius_meters=2, lat_col='latitude', lon_col='longitude'):
    # Load the CSV file
    df = pd.read_csv(csv_file)
    
    # Ensure the columns exist
    if lat_col not in df.columns or lon_col not in df.columns:
        raise ValueError(f"CSV must contain '{lat_col}' and '{lon_col}' columns.")
    
    # Convert coordinates to numeric values (handle potential non-numeric data)
    df[lat_col] = pd.to_numeric(df[lat_col], errors='coerce')
    df[lon_col] = pd.to_numeric(df[lon_col], errors='coerce')
    df = df.dropna(subset=[lat_col, lon_col])  # Remove rows with invalid coordinates
    
    # Compute distances
    target_point = (target_lat, target_lon)
    df['distance'] = df.apply(lambda row: geodesic(target_point, (row[lat_col], row[lon_col])).meters, axis=1)
    
    # Filter rows within the radius
    within_radius = df[df['distance'] <= radius_meters]
    
    return within_radius

if __name__ == "__main__":
    csv_path = input("Enter the path to the CSV file: ")
    target_lat = float(input("Enter the target latitude: "))
    target_lon = float(input("Enter the target longitude: "))
    
    try:
        # Always use a fixed radius of 2 meters
        points_within_radius = find_points_within_radius(csv_path, target_lat, target_lon, radius_meters=2)
        print("Points within 2 meters:")
        print(points_within_radius)
        
        # Calculate the average values for numeric columns
        averages = points_within_radius.mean(numeric_only=True)
        averages['latitude'] = 'Average'  # Add a label for the average row
        averages['longitude'] = ''  # Leave non-numeric columns blank
        
        # Append the averages as a new row to the DataFrame using pd.concat
        averages_df = pd.DataFrame([averages])  # Convert averages to a DataFrame
        points_within_radius = pd.concat([points_within_radius, averages_df], ignore_index=True)
        
        # Save the output to an Excel file
        output_file = "/Users/elizabethskelton/Documents/2024/Masters/Analysis/2mradius/Output/points_within_2m_radius_with_averages.xlsx"
        points_within_radius.to_excel(output_file, index=False)
        
        # Apply bold formatting to the last row (average row)
        wb = load_workbook(output_file)
        ws = wb.active
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)
        wb.save(output_file)
        
        print(f"Results saved to {output_file}")

    except Exception as e:
        print(f"Error: {e}")